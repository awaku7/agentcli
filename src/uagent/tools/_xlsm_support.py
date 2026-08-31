from __future__ import annotations

"""Read-only static analysis for XLSM workbooks and embedded VBA."""

import json
import os
import re
import tempfile
from pathlib import Path
from typing import Any

from .._pip_auto import install_with_status
from .i18n_helper import make_tool_translator

_ = make_tool_translator(__file__)

TOOL_SPEC: dict[str, Any] = {
    "tool_level": 1,
    "tool_genre": "office",
    "type": "function",
    "x_parallel_safe": True,
    "function": {
        "name": "xlsm_analyze",
        "description": _(
            "tool.description",
            default="Statically analyze an XLSM workbook and its embedded VBA without executing macros.",
        ),
        "x_search_terms": _(
            "x_search_terms",
            default=[
                "xlsm analysis",
                "excel macro analysis",
                "vba analysis",
                "analyze xlsm",
            ],
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "input_path": {
                    "type": "string",
                    "description": _(
                        "param.input_path.description",
                        default="Path to the .xlsm file.",
                    ),
                },
                "password": {
                    "type": "string",
                    "description": _(
                        "param.password.description",
                        default="Password for an encrypted workbook (optional).",
                    ),
                },
                "output_path": {
                    "type": "string",
                    "description": _(
                        "param.output_path.description",
                        default="Optional JSON or Markdown output path.",
                    ),
                },
                "output_format": {
                    "type": "string",
                    "enum": ["json", "markdown"],
                    "default": "json",
                    "description": _(
                        "param.output_format.description", default="Output format."
                    ),
                },
                "include_source": {
                    "type": "boolean",
                    "default": True,
                    "description": _(
                        "param.include_source.description",
                        default="Include extracted VBA source code.",
                    ),
                },
                "overwrite": {
                    "type": "boolean",
                    "default": False,
                    "description": _(
                        "param.overwrite.description",
                        default="Allow replacing an existing output file.",
                    ),
                },
            },
            "required": ["input_path"],
            "additionalProperties": False,
        },
    },
}

_SUSPICIOUS = {
    "shell": r"\bShell\s*\(",
    "file_io": r"\b(Open|Kill|Name|FileCopy|Dir|CreateObject)\b",
    "http_or_network": r"(MSXML2|WinHttp|XMLHTTP|InternetExplorer|ADODB\.Stream)",
    "external_workbook": r"\b(Workbooks\.Open|LinkSources|UpdateLink)\b",
    "database": r"(ADODB\.|DAO\.|\.OpenRecordset|ConnectionString)",
}


def _macro_analysis(source: str) -> dict[str, Any]:
    procedures = []
    for match in re.finditer(
        r"^\s*(?:(Public|Private|Friend)\s+)?(Sub|Function|Property\s+(?:Get|Let|Set))\s+([A-Za-z_][\w]*)",
        source,
        re.I | re.M,
    ):
        procedures.append(
            {
                "visibility": match.group(1) or "Public",
                "kind": match.group(2),
                "name": match.group(3),
                "line": source[: match.start()].count("\n") + 1,
            }
        )
    calls = sorted(
        set(re.findall(r"\bCall\s+([A-Za-z_]\w*)|\b([A-Za-z_]\w*)\s*\(", source, re.I))
    )
    call_names = sorted({a or b for a, b in calls if (a or b)})
    findings = []
    for name, pattern in _SUSPICIOUS.items():
        lines = [
            i + 1
            for i, line in enumerate(source.splitlines())
            if re.search(pattern, line, re.I)
        ]
        if lines:
            findings.append({"category": name, "lines": lines})
    return {
        "procedures": procedures,
        "calls": call_names,
        "findings": findings,
        "line_count": len(source.splitlines()),
    }


def _extract_vba(path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    # oletools is an optional, analysis-only dependency. Install it lazily so
    # XLSM analysis works without adding it to the base installation.
    if not install_with_status(
        "oletools", "oletools", verify_submodule="oletools.olevba"
    ):
        raise RuntimeError(
            "VBA extraction requires oletools; automatic installation failed"
        )
    from oletools.olevba import VBA_Parser

    modules: list[dict[str, Any]] = []
    warnings: list[str] = []
    parser = VBA_Parser(str(path))
    try:
        if not parser.detect_vba_macros():
            return modules, warnings
        for filename, stream_path, vba_filename, source in parser.extract_macros():
            source = source or ""
            modules.append(
                {
                    "filename": filename,
                    "stream": stream_path,
                    "module": vba_filename,
                    "analysis": _macro_analysis(source),
                    "source": source,
                }
            )
    except Exception as exc:
        warnings.append(f"VBA extraction failed: {type(exc).__name__}: {exc}")
    finally:
        parser.close()
    return modules, warnings


def _extract_sheets(path: Path) -> list[dict[str, Any]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for worksheet analysis") from exc
    try:
        workbook = openpyxl.load_workbook(
            path, read_only=True, data_only=False, keep_links=True
        )
    except Exception as exc:
        raise RuntimeError(f"could not open workbook: {exc}") from exc
    result = []
    try:
        for sheet in workbook.worksheets:
            formulas = []
            nonempty = 0
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        nonempty += 1
                        if isinstance(cell.value, str) and cell.value.startswith("="):
                            formulas.append(
                                {"cell": cell.coordinate, "formula": cell.value}
                            )
            result.append(
                {
                    "name": sheet.title,
                    "rows": sheet.max_row or 0,
                    "columns": sheet.max_column or 0,
                    "nonempty_cells": nonempty,
                    "formula_count": len(formulas),
                    "formulas": formulas,
                    "merged_ranges": [str(r) for r in sheet.merged_cells.ranges],
                }
            )
    finally:
        workbook.close()
    return result


def _markdown(report: dict[str, Any]) -> str:
    lines = [f"# {report['workbook']['name']}", "", "## Worksheets", ""]
    for sheet in report["workbook"]["sheets"]:
        lines.append(
            f"- **{sheet['name']}**: {sheet['rows']} rows × {sheet['columns']} columns; {sheet['nonempty_cells']} non-empty cells; {sheet['formula_count']} formulas"
        )
    lines += ["", "## VBA modules", ""]
    if not report["vba"]["modules"]:
        lines.append("_No VBA modules detected._")
    for module in report["vba"]["modules"]:
        analysis = module["analysis"]
        lines += [
            f"### {module['module']}",
            f"- Procedures: {len(analysis['procedures'])}",
            f"- Calls: {', '.join(analysis['calls']) or 'none'}",
        ]
        if analysis["findings"]:
            lines.append(
                "- Findings: "
                + ", ".join(
                    f"{x['category']} (lines {','.join(map(str, x['lines']))})"
                    for x in analysis["findings"]
                )
            )
        if report.get("include_source"):
            lines += ["", "```vb", module["source"].rstrip(), "```"]
    if report["warnings"]:
        lines += ["", "## Warnings", ""] + [
            f"- {warning}" for warning in report["warnings"]
        ]
    return "\n".join(lines).rstrip() + "\n"


def _resolve_encrypted(path: Path, password: str | None) -> tuple[Path, Path | None]:
    """Return a readable path and a temporary decrypted path when necessary."""
    try:
        from .._pip_auto import install_with_status

        if not install_with_status("msoffcrypto-tool", "msoffcrypto"):
            return path, None
        import msoffcrypto
    except Exception as exc:
        raise RuntimeError(f"msoffcrypto could not be loaded: {exc}") from exc
    with path.open("rb") as source:
        office = msoffcrypto.OfficeFile(source)
        try:
            encrypted = bool(office.is_encrypted())
        except Exception:
            encrypted = False
        if not encrypted:
            return path, None
        if not password:
            raise ValueError("password is required for encrypted XLSM files")
        office.load_key(password=password)
        fd, temp_name = tempfile.mkstemp(suffix=".xlsm")
        os.close(fd)
        temp_path = Path(temp_name)
        try:
            with temp_path.open("wb") as target:
                office.decrypt(target)
        except Exception:
            temp_path.unlink(missing_ok=True)
            raise
        return temp_path, temp_path


def run_tool(args: dict[str, Any]) -> str:
    args = args or {}
    raw_path = str(args.get("input_path") or "").strip()
    if not raw_path:
        return json.dumps(
            {
                "ok": False,
                "error": _("err.input_path_required", default="input_path is required"),
            },
            ensure_ascii=False,
        )
    path = Path(raw_path)
    if not path.is_file():
        return json.dumps(
            {
                "ok": False,
                "error": _(
                    "err.file_not_found", default="file not found: {path}"
                ).format(path=raw_path),
            },
            ensure_ascii=False,
        )
    if path.suffix.lower() != ".xlsm":
        return json.dumps(
            {
                "ok": False,
                "error": _("err.extension", default="input must be an .xlsm file"),
            },
            ensure_ascii=False,
        )
    temporary: Path | None = None
    try:
        password = str(args.get("password") or "").strip() or None
        readable_path, temporary = _resolve_encrypted(path, password)
        sheets = _extract_sheets(readable_path)
        modules, warnings = _extract_vba(readable_path)
        report: dict[str, Any] = {
            "ok": True,
            "workbook": {"name": path.name, "path": str(path), "sheets": sheets},
            "vba": {"modules": modules},
            "warnings": warnings,
            "include_source": bool(args.get("include_source", True)),
        }
        fmt = str(args.get("output_format") or "json").lower()
        if fmt not in {"json", "markdown"}:
            raise ValueError("output_format must be json or markdown")
        content = (
            _markdown(report)
            if fmt == "markdown"
            else json.dumps(report, ensure_ascii=False, indent=2)
        )
        output = str(args.get("output_path") or "").strip()
        if output:
            destination = Path(output)
            if destination.resolve() == path.resolve():
                raise ValueError(
                    _(
                        "err.input_output_same",
                        default="output_path must not be the input file",
                    )
                )
            if destination.exists() and not args.get("overwrite", False):
                raise FileExistsError(
                    _(
                        "err.output_exists",
                        default="output file already exists: {path}; set overwrite=true to replace it",
                    ).format(path=output)
                )
            destination.parent.mkdir(parents=True, exist_ok=True)
            destination.write_text(content, encoding="utf-8")
        return (
            content
            if not output
            else json.dumps(
                {
                    "ok": True,
                    "output_path": output,
                    "format": fmt,
                    "warnings": warnings,
                },
                ensure_ascii=False,
            )
        )
    except Exception as exc:
        return json.dumps({"ok": False, "error": str(exc)}, ensure_ascii=False)
    finally:
        if temporary is not None:
            temporary.unlink(missing_ok=True)
