from __future__ import annotations

import os
from typing import Any, List, Dict

from .._pip_auto import install_with_status

openpyxl = None
_openpyxl_initialized = False


def _ensure_openpyxl() -> bool:
    global openpyxl, _openpyxl_initialized
    if _openpyxl_initialized:
        return openpyxl is not None
    _openpyxl_initialized = True
    if not install_with_status("openpyxl", "openpyxl"):
        return False
    try:
        import openpyxl as _openpyxl

        openpyxl = _openpyxl
    except Exception:
        openpyxl = None
    return openpyxl is not None


from .i18n_helper import make_tool_translator
from .index_tool_helpers import resolve_index_path

_ = make_tool_translator(__file__)

TOOL_SPEC = {
    "type": "function",
    "tool_genre": "index",
    "function": {
        "name": "excel2idx",
        "description": _(
            "tool.description",
            default=(
                "Parse an Excel workbook (.xlsx/.xlsm) into worksheet summaries and table structures, "
                "returning a numbered index or a specific worksheet section. Use this when reading large Excel files: "
                "call with mode='index' to get the list of worksheets with row/column counts and header previews, then call with "
                "mode='section' and section=N to retrieve the cell data from that worksheet."
            ),
        ),
        "x_search_terms": _(
            "x_search_terms",
            default=[
                "read excel file",
                "excel index",
                "excel parser",
                "worksheet reader",
                "excel section",
                "Read Excel files",
                "Excel index",
                "Sheet index",
                "Extract sheet",
                "Split into sections",
            ],
        ),
        "x_search_terms_en": [
            "read excel file",
            "excel index",
            "excel parser",
            "worksheet reader",
            "excel section",
        ],
        "x_parallel_safe": True,
        "parameters": {
            "type": "object",
            "properties": {
                "path": {
                    "type": "string",
                    "description": _(
                        "param.path.description",
                        default="Path to the Excel (.xlsx/.xlsm) file.",
                    ),
                },
                "mode": {
                    "type": "string",
                    "enum": ["index", "section"],
                    "description": _(
                        "param.mode.description",
                        default=(
                            '"index" returns a numbered table of contents listing worksheets with dimensions and headers. '
                            '"section" returns the table content of a specific worksheet.'
                        ),
                    ),
                },
                "section": {
                    "type": "integer",
                    "description": _(
                        "param.section.description",
                        default=(
                            "Worksheet number to retrieve (1-indexed, used only when mode='section'). "
                            "Get the number from the index output."
                        ),
                    ),
                },
                "max_rows": {
                    "type": "integer",
                    "description": _(
                        "param.max_rows.description",
                        default="Maximum number of rows to return in section mode (default 100).",
                    ),
                },
            },
            "required": ["path", "mode"],
            "additionalProperties": False,
        },
    },
}


class _ExcelIndexBuilder:
    def __init__(self, filepath: str):
        self.filepath = filepath
        self.sheets: List[Dict[str, Any]] = []
        self._parse()

    def _parse(self) -> None:
        if openpyxl is None:
            raise RuntimeError("openpyxl library is required to parse Excel files.")

        wb = openpyxl.load_workbook(self.filepath, read_only=True, data_only=True)
        try:
            for idx, sheet_name in enumerate(wb.sheetnames, 1):
                ws = wb[sheet_name]
                max_r = ws.max_row or 0
                max_c = ws.max_column or 0

                # Preview header from first non-empty row
                headers: List[str] = []
                sample_rows: List[List[str]] = []
                row_count = 0

                for row in ws.iter_rows(values_only=True):
                    row_count += 1
                    row_vals = [str(val) if val is not None else "" for val in row]
                    if any(row_vals):
                        if not headers:
                            headers = [v.replace("\n", " ") for v in row_vals if v]
                        sample_rows.append(row_vals)
                    if len(sample_rows) >= 500:  # Cap preview scan
                        break

                header_preview = (
                    ", ".join(headers[:5]) if headers else "(empty/no headers)"
                )
                if len(headers) > 5:
                    header_preview += ", ..."

                self.sheets.append(
                    {
                        "sheet_num": idx,
                        "name": sheet_name,
                        "max_row": max_r,
                        "max_col": max_c,
                        "header_preview": header_preview,
                    }
                )
        finally:
            wb.close()


def run_tool(args: dict[str, Any]) -> str:
    if not _ensure_openpyxl():
        return _(
            "err.openpyxl_missing",
            default="Error: openpyxl is not installed or could not be imported.",
        )
    path = args.get("path")
    if not path:
        return _("err.path_required", default="Error: 'path' is required.")

    mode = args.get("mode", "index")
    section = args.get("section")
    max_rows = args.get("max_rows", 100)

    try:
        resolved = resolve_index_path(path)
    except Exception as e:
        return _("err.read_error", default="Error reading file: {e}").format(e=e)

    if not os.path.isfile(resolved):
        return _("err.file_not_found", default="Error: File not found: {path}").format(
            path=path
        )

    try:
        builder = _ExcelIndexBuilder(resolved)
    except Exception as e:
        return _("err.parse_error", default="Error parsing Excel file: {e}").format(e=e)

    if not builder.sheets:
        return _("msg.no_sheets", default="(no worksheets found in workbook)")

    if mode == "index":
        toc_lines = []
        for sheet in builder.sheets:
            s_num = sheet["sheet_num"]
            name = sheet["name"]
            r_cnt = sheet["max_row"]
            c_cnt = sheet["max_col"]
            prev = sheet["header_preview"]
            toc_lines.append(
                f"Sheet {s_num:2d}: '{name}' [{r_cnt} rows x {c_cnt} cols] - Headers: {prev}"
            )

        toc = "\n".join(toc_lines)
        return _(
            "msg.index_output",
            default=(
                "Index for: {path}\n---\n{toc}\n---\n"
                "Total sheets: {total}\n"
                "To retrieve a worksheet, call excel2idx with mode='section' and section=N."
            ),
        ).format(path=path, toc=toc, total=len(builder.sheets))

    elif mode == "section":
        if section is None:
            return _(
                "err.section_required",
                default="Error: 'section' (integer) is required when mode='section'.",
            )
        try:
            section_num = int(section)
        except (ValueError, TypeError):
            return _(
                "err.section_invalid", default="Error: 'section' must be an integer."
            )

        if section_num < 1 or section_num > len(builder.sheets):
            return _(
                "err.section_not_found",
                default="Error: Sheet {section_num} not found. Valid range: 1..{last}.",
            ).format(section_num=section_num, last=len(builder.sheets))

        target_sheet = builder.sheets[section_num - 1]
        sheet_name = target_sheet["name"]

        wb = openpyxl.load_workbook(resolved, read_only=True, data_only=True)
        try:
            ws = wb[sheet_name]
            rows_data: List[str] = []
            count = 0
            for row in ws.iter_rows(values_only=True):
                count += 1
                row_vals = [str(val) if val is not None else "" for val in row]
                if any(row_vals):
                    rows_data.append(f"{count:4d} | " + " | ".join(row_vals))
                if count >= max_rows:
                    break

            out_lines = [
                f"=== Sheet {section_num}: '{sheet_name}' ({len(rows_data)} rows displayed) ==="
            ]
            out_lines.extend(rows_data)
            return "\n".join(out_lines)
        finally:
            wb.close()

    else:
        return _(
            "err.invalid_mode",
            default="Error: Invalid mode '{mode}'. Use 'index' or 'section'.",
        ).format(mode=mode)
