from __future__ import annotations

"""Read-only analysis of XLSX/XLSM worksheets, formulas, and VBA."""

import json
import re
from pathlib import Path
from typing import Any

from .i18n_helper import make_tool_translator

_ = make_tool_translator(__file__)

TOOL_SPEC: dict[str, Any] = {
    "tool_level": 1,
    "tool_genre": "office",
    "type": "function",
    "x_parallel_safe": True,
    "function": {
        "name": "spreadsheet_analyze",
        "description": _(
            "tool.description",
            default="Analyze XLSX/XLSM worksheets, formulas, dependencies, and embedded VBA without executing macros.",
        ),
        "x_search_terms": _(
            "x_search_terms",
            default=[
                "spreadsheet analysis",
                "xlsx analysis",
                "xlsm analysis",
                "formula dependencies",
                "vba analysis",
            ],
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "input_path": {
                    "type": "string",
                    "description": _(
                        "param.input_path.description",
                        default="Path to an .xlsx or .xlsm file.",
                    ),
                },
                "output_path": {
                    "type": "string",
                    "description": _(
                        "param.output_path.description",
                        default="Optional JSON or Markdown output path.",
                    ),
                },
                "output_format": {
                    "type": "string",
                    "enum": ["json", "markdown"],
                    "default": "json",
                    "description": _(
                        "param.output_format.description", default="Output format."
                    ),
                },
                "password": {
                    "type": "string",
                    "description": _(
                        "param.password.description",
                        default="Optional password for an encrypted workbook.",
                    ),
                },
                "include_source": {
                    "type": "boolean",
                    "default": True,
                    "description": _(
                        "param.include_source.description",
                        default="Include extracted VBA source for XLSM files.",
                    ),
                },
                "overwrite": {
                    "type": "boolean",
                    "default": False,
                    "description": _(
                        "param.overwrite.description",
                        default="Allow replacing an existing output file.",
                    ),
                },
            },
            "required": ["input_path"],
            "additionalProperties": False,
        },
    },
}

_CELL_REF = re.compile(r"(?:(?:'([^']+)'|([A-Za-z_][\w ]*))!)?\$?([A-Z]{1,3})\$?(\d+)")


def _sheet_data(path: Path) -> tuple[list[dict[str, Any]], dict[str, list[str]]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for worksheet analysis") from exc
    workbook = openpyxl.load_workbook(
        path, read_only=True, data_only=False, keep_links=True
    )
    sheets: list[dict[str, Any]] = []
    dependencies: dict[str, list[str]] = {}
    try:
        for sheet in workbook.worksheets:
            formulas: list[dict[str, Any]] = []
            nonempty = 0
            for row in sheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if value is not None:
                        nonempty += 1
                    if isinstance(value, str) and value.startswith("="):
                        refs = []
                        for match in _CELL_REF.finditer(value):
                            ref_sheet = match.group(1) or match.group(2) or sheet.title
                            refs.append(f"{ref_sheet}!{match.group(3)}{match.group(4)}")
                        key = f"{sheet.title}!{cell.coordinate}"
                        dependencies[key] = sorted(set(refs))
                        formulas.append(
                            {
                                "cell": cell.coordinate,
                                "formula": value,
                                "references": dependencies[key],
                            }
                        )
            sheets.append(
                {
                    "name": sheet.title,
                    "rows": sheet.max_row or 0,
                    "columns": sheet.max_column or 0,
                    "nonempty_cells": nonempty,
                    "formula_count": len(formulas),
                    "formulas": formulas,
                    "merged_ranges": [str(r) for r in sheet.merged_cells.ranges],
                }
            )
    finally:
        workbook.close()
    return sheets, dependencies


def _markdown(report: dict[str, Any]) -> str:
    lines = [f"# {report['workbook']['name']}", "", "## Worksheets", ""]
    for sheet in report["workbook"]["sheets"]:
        lines.append(
            f"- **{sheet['name']}**: {sheet['rows']} rows × {sheet['columns']} columns; {sheet['nonempty_cells']} non-empty cells; {sheet['formula_count']} formulas"
        )
        for formula in sheet["formulas"]:
            lines.append(
                f"  - `{formula['cell']}` = `{formula['formula']}` → {', '.join(formula['references']) or 'no references'}"
            )
    lines += ["", "## Dependencies", ""]
    lines += [
        f"- `{source}` → {', '.join(targets) or 'none'}"
        for source, targets in report["dependencies"].items()
    ] or ["_No formula dependencies detected._"]
    lines += ["", "## VBA", ""]
    lines.append(f"- Modules: {len(report['vba']['modules'])}")
    for module in report["vba"]["modules"]:
        lines.append(
            f"- `{module['module']}`: {len(module['analysis']['procedures'])} procedures; calls: {', '.join(module['analysis']['calls']) or 'none'}"
        )
        if report["include_source"]:
            lines += ["", "```vb", module["source"].rstrip(), "```"]
    return "\n".join(lines).rstrip() + "\n"


def run_tool(args: dict[str, Any]) -> str:
    args = args or {}
    raw = str(args.get("input_path") or "").strip()
    if not raw:
        return json.dumps(
            {
                "ok": False,
                "error": _("err.input_required", default="input_path is required"),
            },
            ensure_ascii=False,
        )
    original = Path(raw)
    if not original.is_file() or original.suffix.lower() not in {".xlsx", ".xlsm"}:
        return json.dumps(
            {
                "ok": False,
                "error": _(
                    "err.extension", default="input must be an .xlsx or .xlsm file"
                ),
            },
            ensure_ascii=False,
        )
    temporary: Path | None = None
    try:
        from ._xlsm_support import _resolve_encrypted, _extract_vba

        readable, temporary = _resolve_encrypted(
            original, str(args.get("password") or "").strip() or None
        )
        sheets, dependencies = _sheet_data(readable)
        modules, warnings = (
            _extract_vba(readable) if original.suffix.lower() == ".xlsm" else ([], [])
        )
        report = {
            "ok": True,
            "workbook": {
                "name": original.name,
                "path": str(original),
                "sheets": sheets,
            },
            "dependencies": dependencies,
            "vba": {"modules": modules},
            "warnings": warnings,
            "include_source": bool(args.get("include_source", True)),
        }
        fmt = str(args.get("output_format") or "json").lower()
        if fmt not in {"json", "markdown"}:
            raise ValueError("output_format must be json or markdown")
        content = (
            _markdown(report)
            if fmt == "markdown"
            else json.dumps(report, ensure_ascii=False, indent=2)
        )
        output = str(args.get("output_path") or "").strip()
        if output:
            destination = Path(output)
            if destination.resolve() == original.resolve():
                raise ValueError(
                    _(
                        "err.same_output",
                        default="output_path must not be the input file",
                    )
                )
            if destination.exists() and not args.get("overwrite", False):
                raise FileExistsError(
                    _(
                        "err.output_exists",
                        default="output file already exists: {path}; set overwrite=true to replace it",
                    ).format(path=output)
                )
            destination.parent.mkdir(parents=True, exist_ok=True)
            destination.write_text(content, encoding="utf-8")
            return json.dumps(
                {
                    "ok": True,
                    "output_path": output,
                    "format": fmt,
                    "warnings": warnings,
                },
                ensure_ascii=False,
            )
        return content
    except Exception as exc:
        return json.dumps({"ok": False, "error": str(exc)}, ensure_ascii=False)
    finally:
        if temporary is not None:
            temporary.unlink(missing_ok=True)
