# tools/excel_ops_tool.py
from __future__ import annotations

import json
import os
from io import BytesIO
from typing import Any

from .i18n_helper import make_tool_translator

try:
    import msoffcrypto
except Exception:  # pragma: no cover
    msoffcrypto = None  # type: ignore[assignment]

_ = make_tool_translator(__file__)

BUSY_LABEL = True

TOOL_SPEC: dict[str, Any] = {
    "tool_level": 1,
    "tool_genre": "office",
    "type": "function",
    "function": {
        "name": "excel_ops",
        "description": _(
            "tool.description",
            default=(
                "Read/write an Excel (.xlsx) file and get sheet names. When writing to an existing file, "
                "a backup with the same name (<file_path>.org / <file_path>.org1 / ...) is created immediately before writing."
            ),
        ),
        "system_prompt": _(
            "tool.system_prompt",
            default="This tool performs the operation described by the tool name 'excel_ops'.",
        ),
        "x_search_terms": _(
            "x_search_terms",
            default=[
                "excel_ops",
                "excel ops",
                "xlsx",
                "spreadsheet",
                "sheet names",
                "excel file",
            ],
        ),
        "x_search_terms_en": [
            "excel_ops",
            "excel ops",
            "xlsx",
            "spreadsheet",
            "sheet names",
            "excel file",
        ],
        "parameters": {
            "type": "object",
            "properties": {
                "action": {
                    "type": "string",
                    "enum": ["read", "write", "get_sheet_names", "keep_only_sheets"],
                    "description": _(
                        "param.action.description",
                        default=(
                            "Operation to perform.\n"
                            "- 'read': read the specified sheet and return JSON.\n"
                            "- 'write': write JSON data to the specified sheet (create file if missing).\n"
                            "- 'get_sheet_names': get a list of sheet names.\n"
                            "- 'keep_only_sheets': create a new workbook (output_path) that keeps only the specified sheets. "
                            "It copies values (not formulas) from the source workbook."
                        ),
                    ),
                },
                "file_path": {
                    "type": "string",
                    "description": _(
                        "param.file_path.description",
                        default="Absolute path to the Excel file.",
                    ),
                },
                "password": {
                    "type": "string",
                    "description": _(
                        "param.password.description",
                        default=(
                            "Optional password for encrypted .xlsx files. If omitted and the file is encrypted, "
                            "the tool will prompt once for a password."
                        ),
                    ),
                },
                "sheet_name": {
                    "type": "string",
                    "description": _(
                        "param.sheet_name.description",
                        default=(
                            "Target sheet name (for read/write). If omitted, read uses the first sheet and write uses 'Sheet1'."
                        ),
                    ),
                },
                "data": {
                    "type": "string",
                    "description": _(
                        "param.data.description",
                        default="Data to write (JSON string).",
                    ),
                },
                "keep_sheets": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": _(
                        "param.keep_sheets.description",
                        default=(
                            "Sheet names to keep (used only for action='keep_only_sheets'). "
                            "If any specified sheet does not exist, the tool errors."
                        ),
                    ),
                },
                "output_path": {
                    "type": "string",
                    "description": _(
                        "param.output_path.description",
                        default=(
                            "Output path for action='keep_only_sheets'. If it already exists, it is overwritten after creating a backup "
                            "(<output_path>.org / <output_path>.org1 / ...)."
                        ),
                    ),
                },
            },
            "required": ["action", "file_path"],
        },
    },
}


def _backup_path(path: str) -> str:
    base = path + ".org"
    if not os.path.exists(base):
        return base
    i = 1
    while True:
        cand = f"{path}.org{i}"
        if not os.path.exists(cand):
            return cand
        i += 1


def _backup_file_if_exists(path: str) -> str | None:
    if not os.path.exists(path):
        return None
    backup = _backup_path(path)
    with open(path, "rb") as fsrc, open(backup, "wb") as fdst:
        fdst.write(fsrc.read())
    return backup


def _prompt_for_password(path: str) -> str | None:
    try:
        from .human_ask_tool import run_tool as human_ask
    except Exception:
        return None

    message = _(
        "prompt.password",
        default="Enter the password for this file:\n{path}",
    ).format(path=path)
    try:
        res_json = human_ask({"message": message, "is_password": True})
        res = json.loads(res_json)
    except Exception:
        return None

    pwd = str(res.get("user_reply") or "").strip()
    return pwd or None


def _load_workbook_with_password(
    file_path: str, password: str | None = None, *, data_only: bool = False
):
    import openpyxl

    if msoffcrypto is not None:
        with open(file_path, "rb") as fin:
            office = msoffcrypto.OfficeFile(fin)
            encrypted = False
            try:
                encrypted = bool(office.is_encrypted())
            except Exception:
                encrypted = False

            if encrypted:
                if not password:
                    password = _prompt_for_password(file_path)
                if not password:
                    raise RuntimeError(
                        "password is required for encrypted workbook files"
                    )

                office.load_key(password=password)
                decrypted = BytesIO()
                office.decrypt(decrypted)
                decrypted.seek(0)
                return openpyxl.load_workbook(decrypted, data_only=data_only)

    return openpyxl.load_workbook(file_path, data_only=data_only)


def run_tool(args: dict[str, Any]) -> str:
    action = args.get("action")
    file_path = str(args.get("file_path", "") or "").strip()
    sheet_name = args.get("sheet_name")
    data_str = args.get("data")
    password = str(args.get("password") or "").strip() or None

    if action not in ("read", "write", "get_sheet_names", "keep_only_sheets"):
        raise ValueError("Invalid action")

    if not file_path:
        raise ValueError("file_path is required")

    # Lazy import openpyxl to keep tool import light.
    import openpyxl

    if action == "get_sheet_names":
        wb = _load_workbook_with_password(file_path, password=password)
        try:
            return json.dumps(wb.sheetnames, ensure_ascii=False)
        finally:
            wb.close()

    if action == "read":
        wb = _load_workbook_with_password(file_path, password=password, data_only=True)
        try:
            target = sheet_name or wb.sheetnames[0]
            ws = wb[target]
            # Return as list of rows (list of values)
            rows: list[list[Any]] = []
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))
            return json.dumps(rows, ensure_ascii=False)
        finally:
            wb.close()

    if action == "keep_only_sheets":
        keep_sheets = args.get("keep_sheets")
        output_path = str(args.get("output_path", "") or "").strip()

        if (
            not isinstance(keep_sheets, list)
            or not keep_sheets
            or not all(isinstance(s, str) and s.strip() for s in keep_sheets)
        ):
            raise ValueError(
                "keep_sheets (non-empty array of sheet names) is required for keep_only_sheets"
            )

        keep_sheets = [str(s).strip() for s in keep_sheets]

        if not output_path:
            raise ValueError("output_path is required for keep_only_sheets")

        # Overwrite allowed; create backup if exists.
        backup = _backup_file_if_exists(output_path)

        # Read values from a recalculated file (call recalc_excel before this tool if needed).
        wb_src = _load_workbook_with_password(
            file_path, password=password, data_only=True
        )
        try:
            missing = [s for s in keep_sheets if s not in wb_src.sheetnames]
            if missing:
                raise ValueError(f"keep_sheets contains non-existent sheets: {missing}")

            wb_out = openpyxl.Workbook()
            try:
                # Remove default sheet
                if wb_out.sheetnames:
                    wb_out.remove(wb_out[wb_out.sheetnames[0]])

                for sname in keep_sheets:
                    ws_src = wb_src[sname]
                    ws_out = wb_out.create_sheet(title=sname)

                    # Best-effort layout copy
                    try:
                        ws_out.sheet_format = ws_src.sheet_format
                    except Exception:
                        pass

                    # Row heights
                    try:
                        for r, dim in ws_src.row_dimensions.items():
                            if dim and dim.height is not None:
                                ws_out.row_dimensions[r].height = dim.height
                    except Exception:
                        pass

                    # Column widths
                    try:
                        for col, dim in ws_src.column_dimensions.items():
                            if dim and dim.width is not None:
                                ws_out.column_dimensions[col].width = dim.width
                    except Exception:
                        pass

                    # Values
                    for row in ws_src.iter_rows(values_only=True):
                        ws_out.append(list(row))

                    # Merged cells (ranges only)
                    try:
                        for rng in list(ws_src.merged_cells.ranges):
                            ws_out.merge_cells(str(rng))
                    except Exception:
                        pass

                os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
                wb_out.save(output_path)
            finally:
                wb_out.close()

            return json.dumps(
                {
                    "ok": True,
                    "file_path": file_path,
                    "output_path": output_path,
                    "backup": backup,
                    "kept_sheets": keep_sheets,
                    "note": "Values-only copy. Call recalc_excel before this tool if you need updated formula results.",
                },
                ensure_ascii=False,
            )
        finally:
            wb_src.close()

    # write
    if os.path.exists(file_path):
        _backup_file_if_exists(file_path)

    if os.path.exists(file_path):
        wb = _load_workbook_with_password(file_path, password=password)
    else:
        wb = openpyxl.Workbook()

    try:
        target = sheet_name or "Sheet1"
        if target in wb.sheetnames:
            ws = wb[target]
            # clear
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet(title=target)

        if data_str is None:
            data_str = "[]"
        data = json.loads(str(data_str))

        if isinstance(data, list) and data and isinstance(data[0], dict):
            # write header
            keys = list(data[0].keys())
            ws.append(keys)
            for obj in data:
                ws.append([obj.get(k) for k in keys])
        elif isinstance(data, list):
            for row in data:
                if isinstance(row, list):
                    ws.append(row)
                else:
                    ws.append([row])
        else:
            ws.append([data])

        os.makedirs(os.path.dirname(file_path) or ".", exist_ok=True)
        wb.save(file_path)
        return file_path
    finally:
        wb.close()
